import openpyxl
from pathlib import Path

path = "/Volumes/data&proj/konkani/outputs/predictions_analysis.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Predictions"]

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers}")

# Find rows with highest WER for 4-gram
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Audio(0), SNR(1), Ref(2), Greedy(3), NoLM(4), 3g(5), 4g(6), G_WER(7), NoLM_WER(8), 3g_WER(9), 4g_WER(10)
    rows.append(row)

# Sort by 4-gram WER descending
worst = sorted([r for r in rows if r[10] is not None], key=lambda x: x[10], reverse=True)[:10]
# Sort by SNR ascending (noisy files)
noisiest = sorted([r for r in rows if r[1] is not None], key=lambda x: x[1], reverse=False)[:10]

print("\n--- WORST PERFORMING (4-GRAM WER) ---")
for r in worst:
    print(f"SNR: {r[1]}dB | WER: {r[10]:.1%} | File: {r[0]}")
    print(f"Ref: {r[2]}")
    print(f"Hyp: {r[6]}\n")

print("\n--- NOISIEST SAMPLES ---")
for r in noisiest:
    print(f"SNR: {r[1]}dB | WER: {r[10]:.1%} | File: {r[0]}")

# Get Average SNR for high WER vs Low WER
high_wer_snr = [r[1] for r in rows if r[10] is not None and r[10] > 0.5]
low_wer_snr = [r[1] for r in rows if r[10] is not None and r[10] < 0.1]

if high_wer_snr:
    print(f"\nAvg SNR for High WER (>50%): {sum(high_wer_snr)/len(high_wer_snr):.1f} dB")
if low_wer_snr:
    print(f"Avg SNR for Low WER (<10%): {sum(low_wer_snr)/len(low_wer_snr):.1f} dB")
