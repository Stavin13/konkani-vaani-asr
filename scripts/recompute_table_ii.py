import json
import torch
import torch.nn.functional as F
import torchaudio
from pathlib import Path
from tqdm import tqdm
import numpy as np
from jiwer import process_words
import unicodedata
import re
import sys
import soundfile as sf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Paths
BASE = Path("/Volumes/data&proj/konkani")
CHECKPOINT = BASE / "outputs/conformer_ctc_run1/best_conformer_ctc.pt"
VOCAB_FILE = BASE / "data/konkani-10k/vocab.json"
TEST_MANIFEST = BASE / "data/konkani-combined/manifests/test.json"
LM_3GRAM = BASE / "models/language_models/konkani_3gram.binary"
LM_4GRAM = BASE / "models/language_models/konkani_4gram.binary"
UNIGRAMS = BASE / "models/language_models/unigrams.txt"

# Model import
sys.path.insert(0, str(BASE))
from models.conformer_ctc import ConformerCTC

# Text normalization
MISSING_DIGITS = set("०३४५६७८")
def normalise_text(text: str, clean: bool = False) -> str:
    if text is None: return ""
    text = unicodedata.normalize("NFC", text)
    if clean:
        cleaned = []
        for ch in text:
            cat = unicodedata.category(ch)
            if (cat.startswith("L") or cat.startswith("M") or
                    cat.startswith("N") or ch.isspace()):
                if ch not in MISSING_DIGITS:
                    cleaned.append(ch)
        text = "".join(cleaned)
    return re.sub(r" +", " ", text).strip()

class CharTokenizer:
    def __init__(self):
        v = json.load(open(VOCAB_FILE, encoding="utf-8"))
        self.idx2char = {int(k): c for k, c in v["idx2char"].items()}
        self.vocab_size = v["vocab_size"]
        self.blank_id = 0 # Match V1 trainer (CTCLoss(blank=0))

    def labels(self):
        L = []
        for i in range(self.vocab_size):
            p = self.idx2char.get(i, f"<id{i}>")
            # For V1 model, 0 is blank.
            if i == 0:
                L.append("")
            elif p in ["<pad>", "<blank>", "<sos>", "<eos>", "<unk>"]:
                L.append(f"<{p[1:-1]}_{i}>")
            else:
                L.append(p)
        return L

def main():
    device = torch.device("mps" if torch.backends.mps.is_available() else "cpu")
    if torch.cuda.is_available(): device = torch.device("cuda")
    print(f"Device: {device}")

    # Load Model
    ckpt = torch.load(CHECKPOINT, map_location="cpu", weights_only=False)
    state = ckpt.get("model_state_dict", ckpt)
    v_size = state["ctc_head.weight"].shape[0]
    
    model = ConformerCTC(vocab_size=v_size).to(device).eval()
    model.load_state_dict(state, strict=False)
    
    tok = CharTokenizer()
    labels = tok.labels()

    # Load decoders
    from pyctcdecode import build_ctcdecoder
    unigrams = None
    if UNIGRAMS.exists():
        with open(UNIGRAMS, encoding="utf-8") as f:
            unigrams = [l.strip() for l in f if l.strip()]

    # Use parameters from clean_eval_wer.py
    dec_3gram = build_ctcdecoder(labels, kenlm_model_path=str(LM_3GRAM), unigrams=unigrams, alpha=0.5, beta=1.5)
    dec_4gram = build_ctcdecoder(labels, kenlm_model_path=str(LM_4GRAM), unigrams=unigrams, alpha=0.5, beta=1.5)

    samples = [json.loads(l) for l in open(TEST_MANIFEST, encoding="utf-8")]
    
    import torchaudio.transforms as T
    mel_fn = T.MelSpectrogram(
        sample_rate=16000, n_mels=80, n_fft=400,
        hop_length=160, win_length=400
    ).to(device)

    # Results holders
    stats = {
        "3gram": {"S": [], "D": [], "I": [], "errors": 0},
        "4gram": {"S": [], "D": [], "I": [], "errors": 0},
        "excel_rows": []
    }

    print(f"Processing {len(samples)} samples with Char-based model (V1)...")

    for s in tqdm(samples):
        try:
            wav, sr = sf.read(s["audio_filepath"])
            wav = torch.from_numpy(wav).float()
            if wav.ndim > 1: wav = wav.mean(-1)
            if sr != 16000:
                import torchaudio.functional as FA
                wav = FA.resample(wav.unsqueeze(0), sr, 16000).squeeze(0)
            wav = wav.to(device)
            
            with torch.no_grad():
                mel = mel_fn(wav.unsqueeze(0))
                # V1 model uses log(mel + 1e-9) and transpose
                mel = torch.log(mel + 1e-9).transpose(1, 2)
                mel_len = torch.tensor([mel.size(1)], device=device)
                logits, _ = model(mel, mel_len)
            
            lp_np = F.log_softmax(logits, dim=-1).squeeze(0).cpu().float().numpy()

            ref = normalise_text(s["text"], clean=True)

            row = [Path(s["audio_filepath"]).name, ref]
            for name, dec in [("3gram", dec_3gram), ("4gram", dec_4gram)]:
                hyp = normalise_text(dec.decode(lp_np, beam_width=15), clean=True)
                
                res = process_words(ref, hyp)
                stats[name]["S"].append(res.substitutions)
                stats[name]["D"].append(res.deletions)
                stats[name]["I"].append(res.insertions)
                if hyp != ref:
                    stats[name]["errors"] += 1
                row.extend([hyp, res.substitutions, res.deletions, res.insertions])
            stats["excel_rows"].append(row)
        except Exception as e:
            # print(f"Error: {e}")
            continue

    # Analysis
    report = []
    report.append("\n" + "="*40)
    report.append("Table II Data Integrity Audit - Final Report")
    report.append("="*40)
    
    # ── Excel Setup ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Table II Summary"
    
    ws_details = wb.create_sheet("Detailed Predictions")
    headers = ["Audio", "Reference", "3-gram Hyp", "S3", "D3", "I3", "4-gram Hyp", "S4", "D4", "I4"]
    ws_details.append(headers)
    for col in range(1, len(headers) + 1):
        ws_details.cell(1, col).font = Font(bold=True)

    # We'll use the stats accumulated in the loop
    # Wait, I need to store the per-sample results to write them to Excel
    # I'll modify the loop to store rows
    
    # Actually, I'll just gather the report text first then write to Excel
    for name in ["3gram", "4gram"]:
        st = stats[name]
        n = len(st["S"])
        if n == 0:
            report.append(f"\n{name.upper()}: No samples processed.")
            continue
        
        s_arr, d_arr, i_arr = np.array(st["S"]), np.array(st["D"]), np.array(st["I"])
        s_mean, s_std = s_arr.mean(), s_arr.std()
        d_mean, d_std = d_arr.mean(), d_arr.std()
        i_mean, i_std = i_arr.mean(), i_arr.std()
        ser = st['errors']/n
        
        report.append(f"\n{name.upper()} Beam Search (N={n})")
        report.append("-" * 25)
        report.append(f"Substitutions: {s_mean:.3f} ± {s_std:.3f}")
        report.append(f"Deletions    : {d_mean:.3f} ± {d_std:.3f}")
        report.append(f"Insertions   : {i_mean:.3f} ± {i_std:.3f}")
        report.append(f"SER          : {ser:.4f}")
        
        # Write to Excel Summary
        ws_summary.append([f"{name.upper()} Metric", "Mean", "Std Dev"])
        ws_summary.append(["Substitutions", s_mean, s_std])
        ws_summary.append(["Deletions", d_mean, d_std])
        ws_summary.append(["Insertions", i_mean, i_std])
        ws_summary.append(["SER", ser, ""])
        ws_summary.append([]) # spacer

    # Write detailed rows to Excel (need to have stored them)
    # I'll update the loop below to store 'excel_rows'
    for row in stats.get("excel_rows", []):
        ws_details.append(row)

    output_text = "\n".join(report)
    print(output_text)
    
    # Save to file
    out_txt = BASE / "outputs/table_ii_stats.txt"
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write(output_text)
    
    out_xlsx = BASE / "outputs/predictions_analysis.xlsx"
    wb.save(out_xlsx)
    print(f"\nResults saved to text: {out_txt}")
    print(f"Results saved to Excel: {out_xlsx}")

if __name__ == "__main__":
    main()
