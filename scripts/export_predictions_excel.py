#!/usr/bin/env python3
"""
Export Predictions to Excel — Konkani ASR
==========================================
Runs all 7,515 test samples through 4 decoding strategies and exports
results to a color-coded Excel file for error analysis.

Columns:
  audio_file | reference | greedy | beam_no_lm | beam_3gram | beam_4gram |
  greedy_wer | beam_no_lm_wer | beam_3gram_wer | beam_4gram_wer

Usage:
    python scripts/export_predictions_excel.py [--count N] [--device mps|cpu]
    python scripts/export_predictions_excel.py --count 500   # quick test run
"""

import json
import re
import sys
import time
import argparse
import unicodedata
from pathlib import Path

import torch
import torch.nn.functional as F
import torchaudio
from tqdm import tqdm
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE          = Path(__file__).resolve().parent.parent
CHECKPOINT    = BASE / "outputs/conformer_ctc_run1/best_conformer_ctc.pt"
VOCAB_FILE    = BASE / "data/konkani-10k/vocab.json"
TEST_MANIFEST = BASE / "data/konkani-combined/manifests/test.json"
LM_3GRAM      = BASE / "models/language_models/konkani_3gram.binary"
LM_4GRAM      = BASE / "models/language_models/konkani_4gram.binary"
UNIGRAMS      = BASE / "models/language_models/unigrams.txt"
OUT_EXCEL     = BASE / "outputs/predictions_analysis.xlsx"

sys.path.insert(0, str(BASE))
from models.conformer_ctc import ConformerCTC

# ── Constants ──────────────────────────────────────────────────────────────────
PUNCT_IN_VOCAB = set("!,-.?'")
MISSING_DIGITS = set("०३४५६७८")

# ── Text normalisation ─────────────────────────────────────────────────────────
def normalise_text(text: str, clean: bool = False) -> str:
    """NFC normalize and optionally strip punctuation."""
    text = unicodedata.normalize("NFC", text)
    if clean:
        cleaned = []
        for ch in text:
            cat = unicodedata.category(ch)
            if (cat.startswith("L") or cat.startswith("M") or
                    cat.startswith("N") or ch.isspace()):
                if ch not in MISSING_DIGITS:
                    cleaned.append(ch)
        text = "".join(cleaned)
    return re.sub(r" +", " ", text).strip()

# ── WER (per-sample) ───────────────────────────────────────────────────────────
def word_edit_dist(ref_words, hyp_words):
    m, n = len(ref_words), len(hyp_words)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[0], i
        for j in range(1, n + 1):
            temp = dp[j]
            if ref_words[i - 1] == hyp_words[j - 1]:
                dp[j] = prev
            else:
                dp[j] = 1 + min(prev, dp[j], dp[j - 1])
            prev = temp
    return dp[n]

def sample_wer(ref: str, hyp: str) -> float:
    ref_w = ref.split()
    hyp_w = hyp.split()
    if not ref_w:
        return 0.0 if not hyp_w else 1.0
    return word_edit_dist(ref_w, hyp_w) / len(ref_w)

# ── Audio processing ───────────────────────────────────────────────────────────
def process_audio(path: str, device: torch.device):
    try:
        wav, sr = torchaudio.load(path)
        if sr != 16000:
            wav = torchaudio.functional.resample(wav, sr, 16000)
        if wav.size(0) > 1:
            wav = wav.mean(0, keepdim=True)
        wav = wav.squeeze(0).to(device)

        mel_fn = torchaudio.transforms.MelSpectrogram(
            sample_rate=16000, n_mels=80, n_fft=400,
            hop_length=160, win_length=400
        ).to(device)
        mel = mel_fn(wav.unsqueeze(0))        # (1, 80, T)
        mel = mel.transpose(1, 2)             # (1, T, 80)
        mel = torch.log(mel + 1e-9)
        mel_len = torch.tensor([(wav.size(0) // 160) + 1], device=device)
        return mel.float(), mel_len
    except Exception:
        return None, None

def compute_snr(path: str):
    """Calculate Signal-to-Noise Ratio using Silero VAD."""
    try:
        wav, sr = torchaudio.load(path)
        if sr != 16000:
            wav = torchaudio.functional.resample(wav, sr, 16000)
        wav = wav.mean(0)
        
        # Load VAD on demand
        vad_model, utils = torch.hub.load('snakers4/silero-vad', 'silero_vad', trust_repo=True, verbose=False)
        get_speech_timestamps = utils[0]
        speech_ts = get_speech_timestamps(wav, vad_model, sampling_rate=16000)
        
        if not speech_ts: return 0.0
            
        speech_mask = torch.zeros_like(wav, dtype=torch.bool)
        for ts in speech_ts:
            speech_mask[ts['start']:ts['end']] = True
            
        s_samples = wav[speech_mask]
        n_samples = wav[~speech_mask]
        
        if len(s_samples) == 0 or len(n_samples) == 0: return 0.0
        s_pow = (s_samples ** 2).mean().item()
        n_pow = (n_samples ** 2).mean().item()
        
        if n_pow <= 1e-10: return 50.0
        return 10 * torch.log10(torch.tensor(s_pow / n_pow)).item()
    except:
        return 0.0

# ── Tokenizer ──────────────────────────────────────────────────────────────────
class CharTokenizer:
    def __init__(self):
        v = json.load(open(VOCAB_FILE, encoding="utf-8"))
        self.idx2char  = {int(k): c for k, c in v["idx2char"].items()}
        self.vocab_size = v["vocab_size"]
        self.blank_id   = 0

    def greedy_decode(self, ids):
        chars, prev = [], -1
        for i in ids:
            if i != self.blank_id and i != prev:
                ch = self.idx2char.get(i, "")
                if not (ch.startswith("<") and ch.endswith(">")):
                    chars.append(ch)
            prev = i
        return "".join(chars).strip()

    def labels(self):
        """pyctcdecode-compatible label list."""
        L = []
        for i in range(self.vocab_size):
            p = self.idx2char.get(i, f"<id{i}>")
            if p in ["<pad>", "<blank>", "<sos>", "<eos>", "<unk>"]:
                L.append("" if i == self.blank_id else f"<{p[1:-1]}_{i}>")
            else:
                L.append(p)
        return L

# ── Excel styling helpers ──────────────────────────────────────────────────────
def wer_fill(wer_val: float) -> PatternFill:
    """Return a color fill based on WER severity."""
    if wer_val == 0:
        return PatternFill("solid", fgColor="C6EFCE")   # green
    elif wer_val <= 0.15:
        return PatternFill("solid", fgColor="FFEB9C")   # yellow
    elif wer_val <= 0.40:
        return PatternFill("solid", fgColor="FFCC99")   # orange
    else:
        return PatternFill("solid", fgColor="FFC7CE")   # red

def pct(v: float) -> str:
    return f"{v * 100:.1f}%"

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Export ASR predictions to Excel")
    parser.add_argument("--count",  type=int, default=None,
                        help="Limit number of samples (default: all)")
    parser.add_argument("--device", default=None,
                        help="Force device: mps|cpu|cuda (default: auto)")
    parser.add_argument("--alpha",  type=float, default=0.5,
                        help="LM weight for beam search (default: 0.5)")
    parser.add_argument("--beta",   type=float, default=1.5,
                        help="Word insertion bonus (default: 1.5)")
    parser.add_argument("--beam",   type=int, default=15,
                        help="Beam width (default: 15)")
    args = parser.parse_args()

    # Device
    if args.device:
        device = torch.device(args.device)
    elif torch.backends.mps.is_available():
        device = torch.device("mps")
    elif torch.cuda.is_available():
        device = torch.device("cuda")
    else:
        device = torch.device("cpu")
    print(f"\nDevice : {device}")

    # Model
    print(f"Loading checkpoint …")
    ckpt   = torch.load(CHECKPOINT, map_location="cpu", weights_only=False)
    state  = ckpt.get("model_state_dict", ckpt)
    v_size = state["ctc_head.weight"].shape[0]
    model  = ConformerCTC(vocab_size=v_size, input_dim=80, d_model=256, num_layers=12)
    model.load_state_dict(state, strict=False)
    model.eval().to(device)
    print(f"Model  : ConformerCTC (vocab_size={v_size})")

    # Tokenizer
    tok = CharTokenizer()
    labels = tok.labels()

    # Build decoders
    from pyctcdecode import build_ctcdecoder

    # Load unigrams if present
    unigrams = None
    if UNIGRAMS.exists():
        with open(UNIGRAMS, encoding="utf-8") as f:
            unigrams = [l.strip() for l in f if l.strip()]
        print(f"Unigrams: {len(unigrams):,} words loaded")
    else:
        print("Unigrams: not found (accuracy may be reduced)")

    print("Building decoders …")
    dec_no_lm  = build_ctcdecoder(labels)
    dec_3gram  = build_ctcdecoder(labels, kenlm_model_path=str(LM_3GRAM),
                                  unigrams=unigrams, alpha=args.alpha, beta=args.beta)
    dec_4gram  = build_ctcdecoder(labels, kenlm_model_path=str(LM_4GRAM),
                                  unigrams=unigrams, alpha=args.alpha, beta=args.beta)

    # Samples
    samples = [json.loads(l) for l in open(TEST_MANIFEST, encoding="utf-8")]
    if args.count:
        samples = samples[:args.count]
    print(f"Samples : {len(samples):,}  |  beam={args.beam}  alpha={args.alpha}  beta={args.beta}")

    # ── Run inference ──────────────────────────────────────────────────────────
    rows = []   # list of dicts

    for s in tqdm(samples, desc="Decoding", unit="sample"):
        mel, mel_len = process_audio(s["audio_filepath"], device)
        if mel is None:
            continue

        with torch.no_grad():
            logits, _ = model(mel, mel_len)

        ref = normalise_text(s["text"], clean=False)

        # 1. Greedy
        ids     = torch.argmax(logits, dim=-1).squeeze(0).tolist()
        greedy  = normalise_text(tok.greedy_decode(ids), clean=False)

        # 2–4. Beam variants (pyctcdecode native)
        lp = F.log_softmax(logits, dim=-1).squeeze(0).cpu().float().numpy()
        b_no_lm  = normalise_text(dec_no_lm.decode(lp,  beam_width=args.beam), clean=False)
        b_3gram  = normalise_text(dec_3gram.decode(lp,  beam_width=args.beam), clean=False)
        b_4gram  = normalise_text(dec_4gram.decode(lp,  beam_width=args.beam), clean=False)

        # WER per sample (clean mode for fair scoring)
        ref_c    = normalise_text(ref,     clean=True)
        # SNR
        snr_val = compute_snr(s["audio_filepath"])

        rows.append({
            "audio":          Path(s["audio_filepath"]).name,
            "snr":            round(snr_val, 1),
            "reference":      ref,
            "greedy":         greedy,
            "beam_no_lm":     b_no_lm,
            "beam_3gram":     b_3gram,
            "beam_4gram":     b_4gram,
            "greedy_wer":     sample_wer(ref_c, normalise_text(greedy,  clean=True)),
            "no_lm_wer":      sample_wer(ref_c, normalise_text(b_no_lm, clean=True)),
            "gram3_wer":      sample_wer(ref_c, normalise_text(b_3gram, clean=True)),
            "gram4_wer":      sample_wer(ref_c, normalise_text(b_4gram, clean=True)),
        })

    print(f"\nRows collected: {len(rows):,}")

    # ── Build Excel ────────────────────────────────────────────────────────────
    print("Writing Excel …")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Header style
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER   = Border(
        bottom=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin",  color="AAAAAA"),
    )

    HEADERS = [
        "Audio File", "SNR (dB)", "Reference",
        "Greedy", "Beam (no LM)", "Beam 3-gram", "Beam 4-gram",
        "Greedy WER", "No LM WER", "3gram WER", "4gram WER",
    ]
    WIDTHS   = [28, 10, 48, 48, 48, 48, 48, 12, 12, 12, 12]

    for col_i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for row_i, r in enumerate(rows, start=2):
        vals = [
            r["audio"], r["snr"], r["reference"],
            r["greedy"], r["beam_no_lm"], r["beam_3gram"], r["beam_4gram"],
            r["greedy_wer"], r["no_lm_wer"], r["gram3_wer"], r["gram4_wer"],
        ]
        wer_cols = {8: r["greedy_wer"], 9: r["no_lm_wer"],
                    10: r["gram3_wer"],  11: r["gram4_wer"]}

        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i)
            cell.border    = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if col_i in wer_cols:
                cell.value      = val                    # float for sorting
                cell.number_format = "0.0%"
                cell.fill       = wer_fill(val)
                cell.font       = Font(bold=(val > 0.5))
            else:
                cell.value = val
                cell.font  = Font(size=10)

    # Auto-filter on row 1 so user can sort/filter in Excel
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    n = len(rows)
    avg = lambda key: sum(r[key] for r in rows) / n if n else 0

    ws2["A1"] = "Metric"
    ws2["B1"] = "Greedy"
    ws2["C1"] = "Beam (no LM)"
    ws2["D1"] = "Beam 3-gram"
    ws2["E1"] = "Beam 4-gram"

    for cell in ws2["1:1"]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Samples",    n, n, n, n),
        ("Avg WER",    avg("greedy_wer"), avg("no_lm_wer"),
                       avg("gram3_wer"),  avg("gram4_wer")),
        ("WER = 0%",   sum(1 for r in rows if r["greedy_wer"] == 0),
                       sum(1 for r in rows if r["no_lm_wer"]  == 0),
                       sum(1 for r in rows if r["gram3_wer"]  == 0),
                       sum(1 for r in rows if r["gram4_wer"]  == 0)),
        ("WER > 80%",  sum(1 for r in rows if r["greedy_wer"] > 0.8),
                       sum(1 for r in rows if r["no_lm_wer"]  > 0.8),
                       sum(1 for r in rows if r["gram3_wer"]  > 0.8),
                       sum(1 for r in rows if r["gram4_wer"]  > 0.8)),
        ("WER > 50%",  sum(1 for r in rows if r["greedy_wer"] > 0.5),
                       sum(1 for r in rows if r["no_lm_wer"]  > 0.5),
                       sum(1 for r in rows if r["gram3_wer"]  > 0.5),
                       sum(1 for r in rows if r["gram4_wer"]  > 0.5)),
    ]

    for ri, (label, *vals) in enumerate(summary_rows, start=2):
        ws2.cell(row=ri, column=1, value=label).font = Font(bold=True)
        for ci, val in enumerate(vals, start=2):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if label == "Avg WER":
                cell.number_format = "0.0%"
                cell.fill = wer_fill(val)

    for col in ["A", "B", "C", "D", "E"]:
        ws2.column_dimensions[col].width = 18

    OUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_EXCEL)
    print(f"\n✅  Saved → {OUT_EXCEL}")
    print(f"    Rows   : {len(rows):,}")
    print(f"    Greedy avg WER   : {avg('greedy_wer') * 100:.2f}%")
    print(f"    No-LM  avg WER   : {avg('no_lm_wer')  * 100:.2f}%")
    print(f"    3-gram avg WER   : {avg('gram3_wer')   * 100:.2f}%")
    print(f"    4-gram avg WER   : {avg('gram4_wer')   * 100:.2f}%")


if __name__ == "__main__":
    main()
